#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Go Namhyeon <gnh1201@gmail.com>
# Convert oracle dump data to mois opendata scheme

def main(args):
    process_rawdata()
    return 0

# 예외적인 데이터를 입력하면 알아서 학습하여 처리합니다
def process_special_token(token):
    result = False

    cases = [
        "CAL_DATE CAL_ CA CA CA CA CAL_STRING",
        "01-JUL-08 2008 07 01 27 1",
        "02-JUL-08 2008 07 02 27 1",
        "03-JUL-08 2008 07 03 27 1",
        "04-JUL-08 2008 07 04 27 1",
        "05-JUL-08 2008 07 05 27 1",
        "PHASE_COEF ABX_C",
        "11 B",
        "1 B",
        "1 C",
        "3 A",
        "3 C000000075",
        "1 C000000022",
        "2 C000000040",
        "1 C000000008",
        "1 C000000038",
        "CONNECT_DATE CONNECT_IP",
        "23-JUN-16 1.221.136.212",
        "S C C C C",
        "N N N N N"
        "ROLE_SEQ U REG_DT",
    ]

    token_len = len(token)
    token_words_len = len(token.split(" "))
    token_rate = token_len / token_words_len

    for case in cases:
        case_len = len(token)
        case_words_len = len(case.split(" "))
        case_rate = case_len / case_words_len

        if token_rate == case_rate:
            result = True

    return result

def process_token(token):
    tokens = []
    splited_tokens = token.split(" ")

    if token.startswith("Y ") or token.startswith("N ") or process_special_token(token):
        tokens = splited_tokens
    else:
        tokens.append(token)
    return tokens

def write_to_excel(data, tablename):
    print("=== START PROCESSING EXCEL ===")

    tablenames = tablename.split(".")

    wb = openpyxl.load_workbook('data/target2.xlsx')
    ws = wb.get_sheet_by_name('양식'.decode("utf-8"))
    for row in ws['H{}:P{}'.format(ws.min_row, ws.max_row)]:
        row_num = row[0].row
        col_0_value = row[0].value # target table name
        col_5_value = row[5].value # target column name

        if col_0_value == tablenames[1]:
            for k in data:
                if k == col_5_value:
                    print("matched! (row: " + str(row_num) + ") " + str(k) + ": " + str(data[k]))
                    ws['P{}'.format(row_num)] = str(data[k])
                    #print(ws['P{}'.format(row_num)].value)

    # save to file
    wb.save("data/target2.xlsx")

    return False

def process_lines(lines):
    tablename = ""
    num_col = 0
    #len_cols = []

    # step 1
    for line in lines:
        if line.startswith("@TABLENAME"):
            x = line.split(" ")
            tablename = x[1]

        if line.startswith("--"):
            x = line.split(" ")
            num_col = len(x)
            #for a in x:
            #    len_cols.append(len(a))
            
    # print step 1
    print("=== TABLENAME: " + str(tablename) + " ===")
    print("=== NUM OF COLS: " + str(num_col) + " ===")
    #print("lengths of each column: " + str(len_cols))

    # step 2
    rows = []
    for line in lines:
        if line.startswith("--") or line.startswith("@"):
            continue

        tokens = [x.strip() for x in line.replace("\t", " ").split("  ") if x != ""]
        tokens_afterword = []
        for token in tokens:
            d = process_token(token)
            tokens_afterword.extend(d)

        rows.append(tokens_afterword)
        #print "['" + "', '".join(tokens_afterword) + "']"
    
    # step 3
    data = {}
    if len(rows) > 1:
        for i, v in enumerate(rows[0]):
            if i < len(rows[1]):
                data[v] = rows[1][i]
    
    # step 4
    for k in data:
        print(k + "\t\t" + data[k])

    # step 5: write to excel
    write_to_excel(data, tablename)

    print("")

def process_rawdata():
    f = open("data/raw_output_2.txt", 'r')
    
    lines = []
    is_buf = False

    while True:
        raw_line = f.readline()
        line = raw_line.rstrip()

        # EOF
        if not raw_line:
            break
        elif not line:
            continue

        # if end dump
        if line == "@DUMPEND":
            #print("detected @DUMPEND")
            process_lines(lines)
            lines = [] # reset lines
            is_buf = False

            #break # if debug

        # store temporary lines
        if is_buf == True:
            lines.append(line)

        # if start dump
        if line == "@DUMPSTART":
            #print("detected @DUMPSTART")
            is_buf = True

if __name__ == '__main__':
    import sys
    import openpyxl
    sys.exit(main(sys.argv))
